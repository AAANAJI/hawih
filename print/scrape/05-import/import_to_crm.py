#!/usr/bin/env python3
"""
import_to_crm.py — bulk-import the Hawih print catalog into the Rise CRM, which
is the single source of truth behind print.hawih.com.sa.

Run this LOCALLY (on the machine that has the catalog files + images). It reads:
  - catalog-skeleton.json      (11 categories · 222 services · publish-ready copy)
  - catalog-review.xlsx        (SAR prices: PriceTiers + NoVariants tabs)
  - a local images folder       (one image per service)
transforms them, and POSTs to the CRM's token-gated import API:
  - POST {CRM}/index.php/print_api/import        (categories + items, idempotent)
  - POST {CRM}/index.php/print_api/import_image  (hero image per item)

The CRM then feeds the store automatically (live sync + daily rebuild).

Prices come ONLY from the xlsx (the real Hawih prices). The weej reference
prices (options[].values[].additional_price_sar, price_tiers[].source_reference_sar)
and every reference_* field are NEVER sent. A forbidden-string scan runs first.

Usage (see README-import.md for the full runbook):
  python3 import_to_crm.py --dry-run            # validate + preflight report, no writes
  python3 import_to_crm.py --pilot 5            # push 5 services as hidden (draft)
  python3 import_to_crm.py --go                 # push all services
  python3 import_to_crm.py --go --images        # + upload hero images
  python3 import_to_crm.py --go --images --publish   # push + images + visible in store

Auth is your CRM admin login (email + password), sent over HTTPS as Basic auth.
No server/DB setup needed.

Config via env or a .env file next to this script (see .env.example):
  CRM_BASE_URL, CRM_ADMIN_EMAIL, CRM_ADMIN_PASSWORD, SKELETON_PATH, XLSX_PATH, IMAGES_DIR

Dependencies:  pip install requests openpyxl
"""

import argparse
import csv
import getpass
import json
import os
import re
import sys
import time
from pathlib import Path

try:
    import requests
except ImportError:
    sys.exit("Missing dependency: pip install requests openpyxl")

# ----------------------------------------------------------------------------- config

HERE = Path(__file__).resolve().parent
OUT_DIR = HERE / "output"


def load_dotenv():
    """Minimal .env loader (no dependency). Values already in the environment win."""
    envfile = HERE / ".env"
    if envfile.exists():
        for line in envfile.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            k, v = k.strip(), v.strip().strip('"').strip("'")
            os.environ.setdefault(k, v)


def cfg(name, default=None, required=False):
    val = os.environ.get(name, default)
    if required and not val:
        sys.exit(f"Missing required config: {name} (set it in .env or the environment)")
    return val


# ----------------------------------------------------------------- forbidden-string scan

FORBIDDEN = [
    r"weej",
    r"ويج",
    r"مطبعة ويج",
    r"1d8471",
    r"9665\d{8}",
    r"\+9665\d{8}",
    r"\b05\d{8}\b",
]


def forbidden_scan(services):
    """Return a list of (service_id, pattern) leaks in the fields we will publish."""
    leaks = []
    publish_keys = (
        "name_ar", "name_en", "slug_ar", "slug_en",
        "description_ar", "description_en", "tags",
    )
    for svc in services:
        blob_parts = [str(svc.get(k, "")) for k in publish_keys]
        for opt in svc.get("options", []) or []:
            blob_parts.append(str(opt.get("name", "")))
            for v in opt.get("values", []) or []:
                blob_parts.append(str(v.get("label", "")))
        blob = " ".join(blob_parts)
        for pat in FORBIDDEN:
            if re.search(pat, blob, re.IGNORECASE):
                leaks.append((svc.get("id", "?"), pat))
    return leaks


# ---------------------------------------------------------------------------- price load

def _norm(s):
    # Unicode-aware: lowercase and drop whitespace/punctuation/underscores, but
    # KEEP letters+digits of any script (Arabic included) so Arabic option/label
    # keys stay distinct. \w is Unicode by default in Python 3.
    return re.sub(r"[\W_]+", "", str(s or "").strip().lower(), flags=re.UNICODE)


def load_prices(xlsx_path):
    """
    Read the xlsx and return per-service Hawih prices from the (real) price_sar column.
      {service_id: {"tiers": [(option, label, price_float)], "flat": price_or_None}}
    Header matching is case/spacing-insensitive. source_reference_sar is ignored.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("Missing dependency: pip install openpyxl")

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    prices = {}

    def sheet_by_name(*wanted):
        wanted_n = [_norm(w) for w in wanted]
        for name in wb.sheetnames:
            if _norm(name) in wanted_n:
                return wb[name]
        return None

    def header_index(ws):
        rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
        header = next(rows, ()) or ()
        return {_norm(h): i for i, h in enumerate(header) if h is not None}

    def cell(row, idx, key):
        i = idx.get(_norm(key))
        return row[i] if (i is not None and i < len(row)) else None

    def to_price(v):
        if v is None or v == "":
            return None
        try:
            return round(float(str(v).replace(",", "").strip()), 2)
        except (ValueError, TypeError):
            return None

    # PriceTiers: service_id, tier_option, tier_label, price_sar
    ws = sheet_by_name("PriceTiers", "Price Tiers", "PriceTier")
    if ws is not None:
        idx = header_index(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            sid = cell(row, idx, "service_id")
            if not sid:
                continue
            sid = str(sid).strip()
            price = to_price(cell(row, idx, "price_sar"))
            if price is None:
                continue
            option = str(cell(row, idx, "tier_option") or "").strip()
            label = str(cell(row, idx, "tier_label") or "").strip()
            prices.setdefault(sid, {"tiers": [], "flat": None})["tiers"].append((option, label, price))

    # NoVariants: service_id, price_sar (flat)
    ws = sheet_by_name("NoVariants", "No Variants", "Flat")
    if ws is not None:
        idx = header_index(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            sid = cell(row, idx, "service_id")
            if not sid:
                continue
            sid = str(sid).strip()
            price = to_price(cell(row, idx, "price_sar"))
            if price is not None:
                prices.setdefault(sid, {"tiers": [], "flat": None})["flat"] = price

    wb.close()
    return prices


# --------------------------------------------------------------------- build the payload

# Option types that are NOT priced radio choices.
UPLOAD_TYPES = {"file-upload", "image"}
FREE_TEXT_TYPES = {"text", "textarea", "note"}
QTY_HINTS = ("كمية", "الكمية", "عدد", "quantity")


def price_blocks_by_name(tiers):
    """
    Segment the ordered tier rows into consecutive same-option BLOCKS and group
    them by normalized option name, preserving order:
      {name_norm: [{label_norm: price}, {label_norm: price}, ...]}

    A single product often lists the SAME option name more than once — one
    size-specific quantity table per size (e.g. four "كمية استيكر الغطاء" blocks
    for cover sizes 20/25/30/34, each pricier than the last). Keying prices by
    (name, label) alone COLLIDES across those blocks and keeps only the last
    one, so every size ends up charged the most expensive size's price. Blocks
    preserve each table separately; map_options assigns the k-th option group of
    a name its k-th block.
    """
    # A new block starts when the option name changes OR the label sequence
    # restarts (a label already seen in the current block reappears) — the only
    # way to separate consecutive same-named tables ("كمية…" 15/20/… then 15/20/…
    # again), which share both name and labels.
    blocks = []              # [(name_norm, {label_norm: price})], in sheet order
    cur_name = object()      # sentinel that never equals a real key
    cur = None
    for (o, l, p) in tiers:
        nk = _norm(o)
        lk = _norm(l)
        if nk != cur_name or (cur is not None and lk in cur):
            cur = {}
            blocks.append((nk, cur))
            cur_name = nk
        cur[lk] = p
    by_name = {}
    for nk, mp in blocks:
        by_name.setdefault(nk, []).append(mp)
    return by_name


def map_options(service, blocks_by_name, base):
    """
    Build the store's print_options from the skeleton options, pricing each value
    from the xlsx tiers (NEVER from the scraped additional_price_sar).
      blocks_by_name: {name_norm: [ {label_norm: price}, ... ]} — one block per
      occurrence of the option name, in order (see price_blocks_by_name). The
      k-th option group named X takes the k-th block of X, so each size-specific
      quantity table prices its OWN group instead of all sharing the last one.
    Repeated groups are kept (NOT collapsed): the store hides all but the one
    matching the chosen size (conditional-option linking), so each size shows —
    and is charged — its own price.
    Returns (options_list, requires_artwork_bool).
    """
    requires_artwork = bool((service.get("specs") or {}).get("accepted_files"))
    options = []
    occ = {}  # name_norm -> option groups of this name already emitted
    for opt in service.get("options", []) or []:
        otype = str(opt.get("type", "single-option")).strip().lower()
        name = str(opt.get("name", "")).strip()
        if otype in UPLOAD_TYPES:
            requires_artwork = True
            continue
        if otype in FREE_TEXT_TYPES:
            continue  # personalization inputs — not modelled as priced options
        values = opt.get("values", []) or []
        if not values:
            continue
        # This group's own price block: the k-th group named X uses the k-th block.
        nk = _norm(name)
        idx = occ.get(nk, 0)
        occ[nk] = idx + 1
        blist = blocks_by_name.get(nk)
        block = (blist[idx] if idx < len(blist) else blist[-1]) if blist else None
        is_qty = any(h in name for h in QTY_HINTS)
        out_values = []
        for v in values:
            label = str(v.get("label", "")).strip()
            if label == "":
                continue
            tier_price = block.get(_norm(label)) if block else None
            delta = round(tier_price - base, 2) if (tier_price is not None and base is not None) else 0.0
            if delta < 0:
                delta = 0.0
            out_values.append({"label_ar": label, "label_en": label, "price_delta": delta})
        if out_values:
            options.append({
                "name_ar": name,
                "name_en": name,
                "type": "tier" if is_qty else "select",
                "values": out_values,
            })
    return options, requires_artwork


def build_item(service, price_row):
    """
    Turn a skeleton service + its xlsx prices into an import item dict, or
    ('blocked', reason) when it cannot be priced.
    """
    tiers = (price_row or {}).get("tiers", [])
    flat = (price_row or {}).get("flat")

    tier_prices = [p for (_o, _l, p) in tiers if p is not None]
    if flat is not None:
        base = flat
    elif tier_prices:
        base = min(tier_prices)
    else:
        base_field = service.get("price_sar")
        base = float(base_field) if base_field not in (None, "") else None

    if base is None:
        return ("blocked", "no price (fill catalog-review.xlsx)")

    blocks_by_name = price_blocks_by_name(tiers)
    options, requires_artwork = map_options(service, blocks_by_name, base)

    category = service.get("category_id", "")  # resolved to name_ar by the caller

    item = {
        "slug": service.get("slug_en") or service.get("id"),
        "category": category,          # placeholder; caller replaces with category name_ar
        "name_ar": service.get("name_ar", ""),
        "name_en": service.get("name_en", ""),
        "description_ar": service.get("description_ar", ""),
        "description_en": service.get("description_en", ""),
        "rate": base,
        "unit_type": (service.get("specs") or {}).get("unit_type", "") or "",
        "requires_artwork": requires_artwork,
        "options": options,
        "hidden": True,   # default draft; --publish flips this
    }
    return item


def build_categories(skeleton):
    cats = []
    for i, c in enumerate(skeleton.get("categories", [])):
        cats.append({
            "name_ar": c.get("name_ar", ""),
            "name_en": c.get("name_en", ""),
            "slug": c.get("slug_en", "") or c.get("id", ""),
            "sort": c.get("sort_order", i),
        })
    return cats


# ------------------------------------------------------------------------------- HTTP

def api_base(base_url):
    base = base_url.rstrip("/")
    if not base.endswith("index.php"):
        base = base + "/index.php"
    return base + "/print_api"


def post_import(session, url_base, categories, items):
    return session.post(
        url_base + "/import",
        headers={"Content-Type": "application/json"},
        data=json.dumps({"categories": categories, "items": items}, ensure_ascii=False).encode("utf-8"),
        timeout=60,
    )


def post_image(session, url_base, slug, image_path):
    with open(image_path, "rb") as fh:
        return session.post(
            url_base + "/import_image",
            data={"slug": slug},
            files={"image": (Path(image_path).name, fh)},
            timeout=120,
        )


IMAGE_EXTS = ("png", "webp", "jpg", "jpeg", "gif")
_PREFER = ("hero", "og", "main", "cover", "1", "01")


def build_image_index(images_dir):
    """
    Walk IMAGES_DIR once and index every image by folder name AND by file stem
    (lowercased), so lookups work no matter how the folders are laid out:
      generated/<service_id>/hero.png  ·  reference/<product_id>/og.jpg  ·
      <slug>.jpg  ·  <product_id>/gallery_1.jpg ...
    When a folder holds several images, the hero/og/first one is chosen.
    Returns (index: {key -> path}, total_images_seen).
    """
    index, total = {}, 0
    if not images_dir or not Path(images_dir).is_dir():
        return index, 0
    for root, _dirs, files in os.walk(images_dir):
        imgs = [f for f in files if f.rsplit(".", 1)[-1].lower() in IMAGE_EXTS]
        if not imgs:
            continue
        total += len(imgs)
        low = {f.lower(): f for f in imgs}
        best = None
        for p in _PREFER:
            for ext in IMAGE_EXTS:
                if f"{p}.{ext}" in low:
                    best = os.path.join(root, low[f"{p}.{ext}"])
                    break
            if best:
                break
        if not best:
            best = os.path.join(root, sorted(imgs)[0])
        parent = Path(root).name.lower()
        if parent:
            index.setdefault(parent, best)          # folder-name match → best image in it
        for f in imgs:
            index.setdefault(f.rsplit(".", 1)[0].lower(), os.path.join(root, f))  # file-stem match
    return index, total


def find_image(index, service):
    """Match a service to an indexed image by id / slug / legacy product id."""
    for k in (service.get("id"), service.get("slug_en"),
              service.get("reference_product_id"), service.get("reference_slug")):
        if k and str(k).strip().lower() in index:
            return index[str(k).strip().lower()]
    return None


# -------------------------------------------------------------------------------- main

def main():
    load_dotenv()
    ap = argparse.ArgumentParser(description="Bulk-import the Hawih print catalog into the CRM.")
    mode = ap.add_mutually_exclusive_group()
    mode.add_argument("--dry-run", action="store_true", help="validate + write preflight report, no server writes (default)")
    mode.add_argument("--pilot", type=int, metavar="N", help="push the first N services (as hidden drafts)")
    mode.add_argument("--go", action="store_true", help="push ALL services")
    ap.add_argument("--images", action="store_true", help="also upload hero images")
    ap.add_argument("--publish", action="store_true", help="make imported items visible in the store (default: hidden/draft)")
    ap.add_argument("--chunk", type=int, default=25, help="items per import request (default 25)")
    args = ap.parse_args()

    skeleton_path = cfg("SKELETON_PATH", str(HERE / "catalog-skeleton.json"))
    xlsx_path = cfg("XLSX_PATH", str(HERE / "catalog-review.xlsx"))
    images_dir = cfg("IMAGES_DIR", "")
    base_url = cfg("CRM_BASE_URL", "https://crm.hawih.com.sa")
    admin_email = cfg("CRM_ADMIN_EMAIL", "")
    admin_password = cfg("CRM_ADMIN_PASSWORD", "")
    token = cfg("PRINT_IMPORT_TOKEN", "")  # optional alternative to admin login

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    if not Path(skeleton_path).exists():
        sys.exit(f"Skeleton not found: {skeleton_path}")
    skeleton = json.loads(Path(skeleton_path).read_text(encoding="utf-8"))
    services = skeleton.get("services", [])
    cats_by_id = {c.get("id"): c for c in skeleton.get("categories", [])}

    # 1) safety scan
    leaks = forbidden_scan(services)
    if leaks:
        print(f"ABORT — {len(leaks)} forbidden-string leaks in publishable fields, e.g. {leaks[:5]}")
        sys.exit(2)
    print(f"Safety scan clean ({len(services)} services).")

    # 2) prices
    prices = load_prices(xlsx_path) if Path(xlsx_path).exists() else {}
    if not prices:
        print(f"WARNING: no prices loaded from {xlsx_path} — items may be blocked.")

    # 3) build items
    items, blocked = [], []
    for svc in services:
        built = build_item(svc, prices.get(str(svc.get("id"))))
        if isinstance(built, tuple) and built[0] == "blocked":
            blocked.append((svc.get("id"), svc.get("name_ar", ""), built[1]))
            continue
        # resolve category id -> Arabic title (CRM identity)
        cat = cats_by_id.get(svc.get("category_id"))
        built["category"] = (cat or {}).get("name_ar", "")
        built["hidden"] = not args.publish
        if not built["category"]:
            blocked.append((svc.get("id"), svc.get("name_ar", ""), "unknown category"))
            continue
        items.append((svc, built))

    categories = build_categories(skeleton)

    # 4) preflight report
    (OUT_DIR / "preflight-report.md").write_text(
        "# Preflight report\n\n"
        f"- Categories: {len(categories)}\n"
        f"- Services in skeleton: {len(services)}\n"
        f"- Importable items: {len(items)}\n"
        f"- Blocked: {len(blocked)}\n"
        f"- Prices loaded for: {len(prices)} services\n"
        f"- Images dir: {images_dir or '(none)'}\n"
        f"- Publish (visible): {args.publish}\n",
        encoding="utf-8",
    )
    if blocked:
        with open(OUT_DIR / "blocked.csv", "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["service_id", "name_ar", "reason"])
            w.writerows(blocked)
    print(f"Preflight: {len(items)} importable, {len(blocked)} blocked, {len(categories)} categories.")
    print(f"  report → {OUT_DIR/'preflight-report.md'}" + (f" · blocked → {OUT_DIR/'blocked.csv'}" if blocked else ""))

    if args.dry_run or (not args.go and args.pilot is None):
        print("\nDry run — no server writes. Re-run with --pilot N or --go to push.")
        # emit the exact payload for inspection
        (OUT_DIR / "payload-preview.json").write_text(
            json.dumps({"categories": categories, "items": [it for _s, it in items[:10]]}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"  first-10 payload preview → {OUT_DIR/'payload-preview.json'}")
        return

    # 5) auth — your CRM admin login (HTTP Basic), or an optional shared token.
    # If the email is set but the password isn't, prompt for it (no need to put
    # the password in .env or juggle shell env vars).
    if not token and admin_email and not admin_password:
        try:
            admin_password = getpass.getpass(f"CRM password for {admin_email}: ")
        except (EOFError, KeyboardInterrupt):
            admin_password = ""
    session = requests.Session()
    if token:
        session.headers["Authorization"] = f"Bearer {token}"
    elif admin_email and admin_password:
        session.auth = (admin_email, admin_password)  # sent as HTTPS Basic auth
    else:
        sys.exit("Set CRM_ADMIN_EMAIL (and CRM_ADMIN_PASSWORD, or type it when prompted) in .env — or PRINT_IMPORT_TOKEN.")
    url_base = api_base(base_url)

    # select scope
    scope = items if args.go else items[: max(0, args.pilot)]
    print(f"\nPushing {len(scope)} items (publish={args.publish}) to {base_url} …")

    ledger, publog = {}, open(OUT_DIR / "publish-log.jsonl", "a", encoding="utf-8")

    # 5a) categories first
    r = post_import(session, url_base, categories, [])
    publog.write(json.dumps({"phase": "categories", "status": r.status_code, "body": _safe(r)}, ensure_ascii=False) + "\n")
    if r.status_code != 200:
        sys.exit(f"Category import failed [{r.status_code}]: {r.text[:300]}")
    print(f"  categories: {r.json().get('categories')}")

    # 5b) items in chunks
    total_created = total_updated = total_failed = 0
    for i in range(0, len(scope), args.chunk):
        chunk = scope[i : i + args.chunk]
        r = post_import(session, url_base, [], [it for _s, it in chunk])
        publog.write(json.dumps({"phase": "items", "range": [i, i + len(chunk)], "status": r.status_code, "body": _safe(r)}, ensure_ascii=False) + "\n")
        if r.status_code != 200:
            print(f"  chunk {i}-{i+len(chunk)} FAILED [{r.status_code}]: {r.text[:200]}")
            continue
        body = r.json()
        total_created += body["items"]["created"]
        total_updated += body["items"]["updated"]
        total_failed += body["items"]["failed"]
        for res in body.get("results", []):
            if res.get("id"):
                ledger[res["slug"]] = {"store_item_id": res["id"], "status": res["status"]}
        print(f"  items {i}-{i+len(chunk)}: +{body['items']['created']} created, ~{body['items']['updated']} updated, !{body['items']['failed']} failed")
        time.sleep(1.0)

    print(f"Items done: {total_created} created, {total_updated} updated, {total_failed} failed.")

    # 6) images
    if args.images:
        img_index, img_total = build_image_index(images_dir)
        print(f"\nUploading images … (scanned {img_total} image files under: {images_dir or '(IMAGES_DIR not set)'})")
        if img_total == 0:
            print("  ⚠ No image files found there. Point IMAGES_DIR in .env at the folder that")
            print("    actually contains your product images (it's scanned recursively), then re-run")
            print("    with --images. Nothing else needs to change.")
        img_ok = img_miss = img_err = 0
        missing = []
        for svc, it in scope:
            path = find_image(img_index, svc)
            if not path:
                img_miss += 1
                missing.append(svc.get("id"))
                continue
            try:
                r = post_image(session, url_base, it["slug"], path)
                if r.status_code == 200 and r.json().get("success"):
                    img_ok += 1
                    ledger.setdefault(it["slug"], {})["image"] = r.json().get("url")
                else:
                    img_err += 1
                    publog.write(json.dumps({"phase": "image", "slug": it["slug"], "status": r.status_code, "body": _safe(r)}, ensure_ascii=False) + "\n")
            except Exception as e:  # noqa: BLE001
                img_err += 1
                publog.write(json.dumps({"phase": "image", "slug": it["slug"], "error": str(e)}, ensure_ascii=False) + "\n")
            time.sleep(0.4)
        print(f"Images: {img_ok} uploaded, {img_miss} missing (no match), {img_err} errors.")
        if missing and img_total:
            print(f"  no image matched for e.g.: {missing[:8]}")
            print("  (these are matched by service id / slug / legacy product id — from a folder")
            print("   name or a file name. Tell me one of your image paths and I'll map it.)")

    publog.close()
    (OUT_DIR / "sync-ledger.json").write_text(json.dumps(ledger, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nLedger → {OUT_DIR/'sync-ledger.json'} · log → {OUT_DIR/'publish-log.jsonl'}")
    if not args.publish:
        print("Items imported as HIDDEN (draft). Review in the CRM, then re-run with --publish to make them live.")
    print("The store reflects changes live for existing pages; new product pages appear on the next rebuild "
          "(nightly, or trigger the 'Deploy print store' workflow now).")


def _safe(resp):
    try:
        return resp.json()
    except Exception:  # noqa: BLE001
        return resp.text[:300]


if __name__ == "__main__":
    main()
